from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from .calculos import normalizar_estado_kanban, nombre_mes, semana_mes, criticidad

COLUMN_MAP = {
    "N° OP": "codigo_op",
    "Fecha recepción": "fecha_recepcion",
    "Hora recepción": "hora_recepcion",
    "Cliente": "cliente",
    "Teléfono": "telefono",
    "Tipo pedido": "tipo_pedido",
    "Cantidad unidades": "cantidad_unidades",
    "Miles facturables": "miles_facturables",
    "Precio x 1,000": "precio_mil",
    "Importe estimado": "importe_estimado",
    "Fecha prometida": "fecha_prometida",
    "Fecha programada": "fecha_programada",
    "Fecha entrega real": "fecha_entrega_real",
    "Estado pago": "estado_pago",
    "Materiales": "materiales",
    "Diseño validado": "diseno_validado",
    "Stock crítico": "stock_critico",
    "OC activada": "oc_activada",
    "Traslado min": "traslado_min",
    "Diagnóstico min": "diagnostico_min",
    "Parada/ajuste h": "parada_ajuste_h",
    "Semana": "semana",
    "Puntaje": "puntaje",
    "Prioridad": "prioridad",
    "Estado Kanban": "estado_op",
    "Cumple programa": "cumple_programa",
    "Resultado": "resultado",
    "Observación / relación con indicadores": "observacion",
    "Acción requerida": "accion_requerida",
    "Responsable": "responsable",
    "Alerta WIP": "alerta_wip",
    "Clave": "clave",
}


def _fmt_date(v):
    if v is None or v == "":
        return None
    try:
        dt = pd.to_datetime(v)
        if pd.isna(dt):
            return None
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return str(v)


def _fmt_time(v):
    if v is None or v == "":
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M")
    txt = str(v).strip()
    # Excel can store time as float fraction
    try:
        num = float(txt)
        if 0 <= num < 1:
            mins = round(num * 24 * 60)
            return f"{mins//60:02d}:{mins%60:02d}"
    except Exception:
        pass
    return txt[:5] if len(txt) >= 5 else txt


def encontrar_tabla_base(excel_path: str | Path):
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=45):
        values = [cell.value for cell in row]
        if values and values[0] == "N° OP" and "Cliente" in values:
            header_row = row[0].row
            headers = values
            break
    if header_row is None:
        raise ValueError("No se encontro la tabla 'BASE OP DIGITAL' con encabezado 'N° OP'.")
    return ws, header_row, headers


def cargar_pedidos_mayo(excel_path: str | Path) -> pd.DataFrame:
    ws, header_row, headers = encontrar_tabla_base(excel_path)
    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if not (isinstance(first, str) and first.startswith("OP-TOBE-")):
            # after 20 empty/non OP rows past data, stop
            if r > header_row + 5 and str(first).startswith("9.4"):
                break
            continue
        raw = {}
        for idx, head in enumerate(headers, start=1):
            # La base tiene columnas auxiliares para el Kanban visual. Para evitar sobreescrituras:
            # col. 25 = estado operativo del pedido, col. 29/30 = helper visual del tablero Excel.
            if idx == 25:
                raw["estado_op"] = ws.cell(r, idx).value
                continue
            if idx in {29, 30}:
                continue
            if head is None:
                continue
            key = COLUMN_MAP.get(str(head).strip())
            if key:
                raw[key] = ws.cell(r, idx).value
        # Estado real para la app: se usa el estado operativo de la base, no el helper visual del Excel.
        raw["estado_kanban"] = normalizar_estado_kanban(raw.get("estado_op"))
        raw["clave"] = ws.cell(r, 36).value or ws.cell(r, 29).value or raw.get("clave")
        raw["codigo_op"] = str(raw.get("codigo_op", "")).strip()
        for date_col in ["fecha_recepcion", "fecha_prometida", "fecha_programada", "fecha_entrega_real"]:
            raw[date_col] = _fmt_date(raw.get(date_col))
        raw["hora_recepcion"] = _fmt_time(raw.get("hora_recepcion"))
        raw["mes"] = nombre_mes(raw.get("fecha_recepcion"))
        if not raw.get("semana") or str(raw.get("semana")).lower().startswith("sin"):
            raw["semana"] = semana_mes(raw.get("fecha_recepcion"))
        raw["fuente"] = "Excel To-Be Mayo 2026"
        raw["fecha_creacion"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        raw["fecha_actualizacion"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        raw["criticidad_key"], raw["criticidad"], raw["criticidad_icono"] = criticidad(raw)
        records.append(raw)
    df = pd.DataFrame(records)
    # Forzar orden y tipos numéricos sin cambiar el dato de Excel
    numeric_cols = ["cantidad_unidades", "miles_facturables", "precio_mil", "importe_estimado", "traslado_min", "diagnostico_min", "parada_ajuste_h", "puntaje"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def resumen_calidad(df: pd.DataFrame) -> dict:
    if df.empty:
        return {"registros": 0, "duplicados": 0, "cumplidos": 0, "retrasados": 0, "cumplimiento": 0.0}
    programados = len(df)
    cumplidos = int(df["cumple_programa"].astype(str).str.lower().isin(["sí", "si", "1", "true"]).sum())
    return {
        "registros": int(len(df)),
        "duplicados": int(df["codigo_op"].duplicated().sum()),
        "cumplidos": cumplidos,
        "retrasados": int(programados - cumplidos),
        "cumplimiento": round(cumplidos / programados * 100, 2) if programados else 0,
        "campos_vacios_cliente": int(df["cliente"].isna().sum()),
    }
